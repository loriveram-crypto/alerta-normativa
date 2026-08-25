"""
Mantiene un archivo Excel con todas las novedades encontradas. Si guardas
este archivo dentro de tu carpeta LOCAL de OneDrive (la que ya se
sincroniza sola con la nube en segundo plano), Power BI se puede conectar
directo a OneDrive con tu login normal — sin necesidad de permisos de
Azure ni de tu área de IT.

Cada fila = un item nuevo encontrado. Nunca se borran filas; el archivo
crece con el tiempo. Power BI/Excel se encargan de filtrar/visualizar
por fecha si hace falta.

CONFIGURA LA RUTA: cambia RUTA_EXCEL más abajo para que apunte a tu
carpeta de OneDrive, por ejemplo:
  Windows: C:/Users/TU_USUARIO/OneDrive - Asobancaria/VigilanciaNormativa/normativa.xlsx
  Mac:     /Users/TU_USUARIO/Library/CloudStorage/OneDrive-Asobancaria/VigilanciaNormativa/normativa.xlsx

También puedes definirla con la variable de entorno RUTA_EXCEL en vez de
editar el código, si prefieres no tocar el archivo.
"""
import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook

# ⚠️ AJUSTA ESTA RUTA a tu carpeta de OneDrive antes de correr el script.
# Si defines la variable de entorno RUTA_EXCEL, esa tiene prioridad.
RUTA_EXCEL = Path(os.environ.get(
    
    "RUTA_EXCEL",
    r"C:\Users\priesgo\OneDrive - ASOBANCARIA\Documentos\Automatización\VigilanciaNormativa\normativa.xlsx"
))

ENCABEZADOS = ["Fecha detección", "Entidad", "Fecha del documento", "Título", "Enlace"]


def _obtener_o_crear_libro():
    RUTA_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    if RUTA_EXCEL.exists():
        libro = load_workbook(RUTA_EXCEL)
        hoja = libro.active
    else:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Normativa"
        hoja.append(ENCABEZADOS)
    return libro, hoja


def agregar_filas(items_nuevos: list[dict]):
    if not items_nuevos:
        print("Sin novedades, no se modifica el Excel.")
        return

    libro, hoja = _obtener_o_crear_libro()
    fecha_deteccion = datetime.now().strftime("%Y-%m-%d %H:%M")

    for item in items_nuevos:
        hoja.append([
            fecha_deteccion,
            item["entidad"],
            item.get("fecha", ""),
            item["titulo"],
            item["link"],
        ])

    libro.save(RUTA_EXCEL)
    print(f"Se agregaron {len(items_nuevos)} fila(s) a {RUTA_EXCEL}.")
